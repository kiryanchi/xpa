import io
import os
import shutil
import string
from io import BytesIO

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Alignment, Side, Border
from PIL import Image as PImage
from src.tools.log import Log


class MyStyle(NamedStyle):
    def __init__(self):
        super().__init__(name="MyStyle")
        self.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(style="thin", color="000000")
        self.border = Border(left=bd, top=bd, right=bd, bottom=bd)


class HJ:
    def __init__(self, file):
        self.file = file
        self.root, self.name = os.path.split(file)
        self.wb = None
        self.sheet = None
        self.style = None
        self.height = 0
        self.width = 33.3
        self.images = {}
        self.row = 0

        try:
            self.wb = openpyxl.load_workbook(file)
        except FileNotFoundError as e:
            Log.warning(self, f"{e}")
        else:
            self.sheet = self.wb.active
            self.height = self.sheet.row_dimensions[3].height

            for i in range(3, self.sheet.max_row + 1):
                self._addImages(i)

            Log.debug(self, self.sheet._images)

            for image in self.sheet._images:
                Log.debug(self, f"images: {image}")
                r = image.anchor._from.row + 1
                c = string.ascii_uppercase[image.anchor._from.col]
                cell = f"{c}{r}"

                self.images[cell] = image

            if 'MyStyle' in self.wb.named_styles:
                self.style = "MyStyle"
            else:
                self.style = MyStyle()

    def _addImages(self, row):
        self.images[f"E{row}"] = None
        self.images[f"F{row}"] = None
        self.images[f"G{row}"] = None

    def _deleteImage(self, col, row):
        try:
            self.sheet._images.remove(self.images[f"{col}{row}"])
        except ValueError:
            Log.debug(self, f"{col}{row} 이미지 없음")
        try:
            del self.images[f"{col}{row}"]
        except KeyError:
            Log.debug(self, f"{col}{row} images dict 없음")

    def _deleteImages(self, row):
        self._deleteImage('E', row)
        self._deleteImage('F', row)
        self._deleteImage('G', row)

    def _indexToLine(self, index):
        return index + 3

    def _setLineStyle(self, row, style):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            self.sheet[f"{col}{row}"].style = style

    def addLine(self):
        self.sheet.row_dimensions[self.sheet.max_row + 1].height = self.height
        self._setLineStyle(self.sheet.max_row + 1, self.style)
        self._addImages(self.sheet.max_row)
        Log.debug(self, f"addLine 후 max_row: {self.sheet.max_row}")

    def deleteAllImages(self):
        self.sheet._images.clear()

    def removeLine(self):
        self._deleteImages(self.sheet.max_row)
        self.sheet.row_dimensions[self.sheet.max_row].height = self.sheet.row_dimensions[self.sheet.max_row + 1].height
        self._setLineStyle(self.sheet.max_row, "Normal")
        self.sheet.delete_rows(self.sheet.max_row, 1)
        Log.debug(self, f"removeLine 후 max_row: {self.sheet.max_row}")

    def getLine(self, index):
        return {
                   "작업내역": str(self.sheet[f"B{self._indexToLine(index)}"].value),
                   "선로번호": str(self.sheet[f"C{self._indexToLine(index)}"].value),
                   "전산화번호": str(self.sheet[f"D{self._indexToLine(index)}"].value),
               }, self.images[f"E{self._indexToLine(index)}"], self.images[f"F{self._indexToLine(index)}"], self.images[
                   f"G{self._indexToLine(index)}"]

    def insertImage(self, col, line, imageData):
        if not imageData:
            return
        if self.images[f"{col}{line}"] and self.images[f"{col}{line}"]._data() == imageData:
            print('이미 같은 사진')
            return
        self._deleteImage(col, line)
        bytes_io = self.resizeImage(imageData)
        img = Image(bytes_io)
        img.width, img.height = 272.64, 240
        self.sheet.add_image(img, f"{col}{line}")
        self.images[f"{col}{line}"] = img

    def loadImage(self):
        print(self.images)
        # for image in self.sheet._images:
        #     try:
        #         r = image.anchor._from.row + 1
        #         c = string.ascii_uppercase[image.anchor._from.col]
        #     except AttributeError:
        #         cell = image.anchor
        #     else:
        #         cell = f"{c}{r}"
        #     self.images[cell] = image
        # print(self.images)

    def resizeImage(self, imageData):
        imageData = io.BytesIO(imageData)
        pimg = PImage.open(imageData)
        pimg = pimg.resize((1200, 800))
        bytes_io = BytesIO()
        pimg.save(bytes_io, 'PNG')

        return bytes_io

    def saveLine(self, index, 내역, 명찰, 전경, 근접):
        line = self._indexToLine(index)
        self.sheet[f"A{line}"].value = str(index + 1)

        self.sheet[f"B{line}"].value = 내역['작업내역']
        self.sheet[f"C{line}"].value = 내역['선로번호']
        self.sheet[f"D{line}"].value = 내역['전산화번호']

        if 명찰:
            self.insertImage('E', line, 명찰)
        if 전경:
            self.insertImage('F', line, 전경)
        if 근접:
            self.insertImage('G', line, 근접)

    def save(self, dest=None):
        if dest is None:
            dest = self.file
        self.wb.save(dest)
        print('저장완료')

    @staticmethod
    def createNewFile(dest):
        NEW_FILE = "./static/hj.xlsx"
        shutil.copyfile(NEW_FILE, dest)
        return HJ(dest)
