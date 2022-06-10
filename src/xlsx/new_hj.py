import io
import os
import shutil
import string
from io import BytesIO

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import NamedStyle, Alignment, Side, Border
from PIL import Image as PImage
from src.tools.log import Log


class MyStyle(NamedStyle):
    def __init__(self):
        super().__init__(name="MyStyle")
        self.alignment = Alignment(horizontal="center", vertical="center")
        bd = Side(style="thin", color="000000")
        self.border = Border(left=bd, top=bd, right=bd, bottom=bd)


class HJ:
    def __init__(self, file):
        self.height = 0
        self.file = file
        self.root, self.name = os.path.split(file)

        self.style = None
        self.images = {}

        self.wb = openpyxl.load_workbook(file)
        self.sheet = self.wb.active

        self._initImages()

        if 'MyStyle' in self.wb.named_styles:
            self.style = 'MyStyle'
        else:
            self.style = MyStyle()

    def addImage(self, col, row, imageData):
        if not imageData:
            return

        if self.images[f"{col}{row}"] and self.images[f"{col}{row}"]._data() == imageData:
            return

        self._removeImage(col, row)

    def addLine(self):
        self.sheet.row_dimensions[self.sheet.max_row + 1].height = self.height
        self._setLineStyle(self.sheet.max_row + 1, self.style)
        self._addNoneImages(self.sheet.max_row)

    def _initImages(self):
        for row in range(3, self.sheet.max_row + 1):
            self._addNoneImages(row)

        for image in self.sheet._images:
            cell = f"{string.ascii_uppercase[image.anchor._from.col]}{image.anchor._from.row + 1}"
            self.images[cell] = image

    def _setLineStyle(self, row, style):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
            self.sheet[f"{col}{row}"].style = style

    def _addNoneImages(self, row):
        for col in ['E', 'F', 'G']:
            self.images[f"{col}{row}"] = None

    def _removeImage(self, col, row):
        pass
